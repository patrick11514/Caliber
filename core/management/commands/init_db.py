from django.core.management.base import BaseCommand
from core.models import Gene, GeneVariant, Patient, ClinVarGeneVariant
import glob
import polars as pl
from openpyxl import load_workbook

def sheet_exists(path: str, sheet: str) -> bool:
    wb = load_workbook(path, read_only=True)
    return sheet in wb.sheetnames

def parse_df(df: pl.DataFrame):
    """"
    Parses wanted fields excel data from the given DataFrame, the variable names depends on the format (Finalist/Franklin)
    """
    for row in df.iter_rows(named=True):
        print(row, type(row))
        patient_name = row.get("Name")
        gene_symbol = row.get("Symbol") or row.get("Gene")
        variant = row.get("Variant_class") or row.get("Variation Type")
        chromosome = row.get("Chr")
        position = row.get("Coordinate") or row.get("Start Position")
        dbSNP = row.get("VEP dbSNP ID") or row.get("dbSNP")
        hgvs_c = row.get("HGVSc") or row.get("Transcript")
        hgvs_c += row.get("Nucleotide", "")
        hgvs_p = row.get("HGVSp") or row.get("AA Change")
        category = row.get("Kategorie")
        comment = row.get("Komentář")
        exon = row.get("Exon")
        zygosity = row.get("Genotype") or row.get("Zygosity")
        gnomAD = row.get("gnomAD AF") or row.get("gnomAD (Exome)")
        print(f"Parsed row: patient_name={patient_name}, gene_symbol={gene_symbol}, chromosome={chromosome}, position={position}, dbSNP={dbSNP}, hgvs_c={hgvs_c}, hgvs_p={hgvs_p}, category={category}, comment={comment}, exon={exon}, zygosity={zygosity}, gnomAD={gnomAD}")


class Command(BaseCommand):

    DEFAULT_ROOT_DIR = "."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--root_dir",
            help="Sets the root directory for the imported xlxs files. Default is the current directory.",
            default=self.DEFAULT_ROOT_DIR,
        )
    
    def handle(self, *args: tuple, **options: dict) -> None:
        root_dir = options.get("root_dir") or self.DEFAULT_ROOT_DIR

        for file in glob.iglob(f"{root_dir}/**/*.xls*", recursive=True):
            print(f"Importing data from {file}...")
            df = None        

            if file.endswith(".xlsx") and sheet_exists(file, "default"):
                df = pl.read_excel(file, sheet_name="default")
            else:
                try:
                    df = pl.read_excel(file, sheet_name="Filtr JI")
                except Exception as e:
                    df = pl.read_excel(file)
            
            parse_df(df)
            print(df.head())

        