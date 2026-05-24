from django.core.management.base import BaseCommand
from apps.core.models import Gene, GeneVariant, GeneticReport, Patient, PatientVariant
import glob
import polars as pl
from openpyxl import load_workbook
from django.db import transaction

patient_cache = {}
gene_cache = {}
variant_cache = {}

def sheet_exists(path: str, sheet: str) -> bool:
    wb = load_workbook(path, read_only=True)
    return sheet in wb.sheetnames

def clean_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_int(value: object) -> int | None:
    if value in (None, "", "nan"):
        return None
    return int(value)

def parse_row(row) -> dict:
    cleaned_data = {}
    cleaned_data["patient_name"] = clean_str(row.get("Name"))
    cleaned_data["gene_symbol"] = clean_str(row.get("Symbol") or row.get("Gene"))
    cleaned_data["variant"] = clean_str(row.get("Variant_class") or row.get("Variation Type"))
    cleaned_data["chromosome"] = clean_str(row.get("Chr"))
    cleaned_data["position"] = clean_int(row.get("Coordinate") or row.get("Start Position"))
    cleaned_data["dbSNP"] = clean_str(row.get("VEP dbSNP ID", "") or row.get("dbSNP", ""))
    cleaned_data["hgvs_c"] = clean_str(row.get("HGVSc") or row.get("Transcript"))
    cleaned_data["hgvs_c"] += clean_str(row.get("Nucleotide", ""))
    cleaned_data["hgvs_p"] = clean_str(row.get("HGVSp", "") or row.get("AA Change", ""))
    cleaned_data["category"] = clean_str(row.get("Kategorie"))
    cleaned_data["comment"] = clean_str(row.get("Komentář", ""))
    cleaned_data["exon"] = clean_str(row.get("Exon"))
    cleaned_data["zygosity"] = clean_str(row.get("Genotype") or row.get("Zygosity"))
    cleaned_data["gnomAD"] = clean_str(row.get("gnomAD AF") or row.get("gnomAD (Exome)"))

    return cleaned_data

def persist_row(data: dict, file_name: str):
    if data["patient_name"] not in patient_cache:
        patient, _ = Patient.objects.get_or_create(name=data["patient_name"])
        patient_cache[data["patient_name"]] = patient
    else:
        patient = patient_cache[data["patient_name"]]

    if data["gene_symbol"] not in gene_cache:
        gene, _ = Gene.objects.get_or_create(symbol=data["gene_symbol"])
        gene_cache[data["gene_symbol"]] = gene
    else:
        gene = gene_cache[data["gene_symbol"]]

    if (data["gene_symbol"], data["chromosome"], data["position"], data["variant"], data["hgvs_c"]) not in variant_cache:
        gene_variant, _ = GeneVariant.objects.get_or_create(
            gene=gene,
            chromosome=data["chromosome"],
            position=data["position"],
            variation_type=data["variant"],
            hgvs_c=data["hgvs_c"],
            defaults={
                "hgvs_p": data["hgvs_p"],
                "dbsnp": data["dbSNP"],
            }
        )
        variant_cache[(data["gene_symbol"], data["chromosome"], data["position"], data["variant"], data["hgvs_c"])] = gene_variant
    else:
        gene_variant = variant_cache[(data["gene_symbol"], data["chromosome"], data["position"], data["variant"], data["hgvs_c"])]

    # TODO - ADD date of the file creation as the created_at and updated_at so it matches the date of the report, not the date of the import
    report, _ = GeneticReport.objects.get_or_create(
        patient=patient,
        report_name=f"{file_name}"
    )

    p_v, _ = PatientVariant.objects.get_or_create(
        report=report,
        variant=gene_variant,
        exon=data["exon"],
        gnomAD=data["gnomAD"],
        zygosity=data["zygosity"],
        category=data["category"],
        comment=data["comment"],
    )

def parse_df(df: pl.DataFrame, file_name: str):
    """"
    Parses wanted fields excel data from the given DataFrame, the variable names depends on the format (Finalist/Franklin)
    """
    for row in df.iter_rows(named=True):
        cleaned_data = parse_row(row)
        persist_row(cleaned_data, file_name)
        


class Command(BaseCommand):

    DEFAULT_ROOT_DIR = "."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--root_dir",
            help="Sets the root directory for the imported xlxs files. Default is the current directory.",
            default=self.DEFAULT_ROOT_DIR,
        )
    
    def handle(self, *args: tuple, **options: dict) -> None:
        root_dir = options.get("root_dir") or self.DEFAULT_ROOT_DIR

        for file_name in glob.iglob(f"{root_dir}/**/*.xls*", recursive=True):
            print(f"Importing data from {file_name}...")
            df = None

            if file_name.endswith(".xlsx") and sheet_exists(file_name, "default"):
                df = pl.read_excel(file_name, sheet_name="default")
            else:
                try:
                    df = pl.read_excel(file_name, sheet_name="Filtr JI")
                except Exception as e:
                    df = pl.read_excel(file_name)
            
            with transaction.atomic():
                parse_df(df, file_name)

        